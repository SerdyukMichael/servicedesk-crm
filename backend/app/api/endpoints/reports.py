from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.api.deps import get_db, require_roles
from app.models import Ticket, User
from app.schemas import TicketReportResponse

router = APIRouter()
_ROLES = ("director", "svc_mgr", "admin")

FINAL_STATUSES = {"completed", "closed", "cancelled"}


def _build_report(
    db: Session,
    date_from: date,
    date_to: date,
    engineer_id: Optional[int],
    client_id: Optional[int],
) -> dict:
    q = (
        db.query(Ticket)
        .options(joinedload(Ticket.assignee), joinedload(Ticket.client))
        .filter(
            Ticket.is_deleted.is_(False),
            Ticket.created_at >= datetime.combine(date_from, time.min),
            Ticket.created_at <= datetime.combine(date_to, time.max),
        )
    )
    if engineer_id:
        q = q.filter(Ticket.assigned_to == engineer_id)
    if client_id:
        q = q.filter(Ticket.client_id == client_id)

    rows = q.all()
    total = len(rows)

    by_status: dict[str, int] = {}
    by_type: dict[str, int] = {}
    by_priority: dict[str, int] = {}
    by_engineer: dict[str, int] = {}

    reaction_total = reaction_violated = 0
    resolution_total = resolution_violated = 0
    resolution_times: list[float] = []

    now = datetime.utcnow()

    for t in rows:
        by_status[t.status] = by_status.get(t.status, 0) + 1
        by_type[t.type] = by_type.get(t.type, 0) + 1
        by_priority[t.priority] = by_priority.get(t.priority, 0) + 1

        if t.assigned_to and t.assignee:
            name = t.assignee.full_name
            by_engineer[name] = by_engineer.get(name, 0) + 1

        # SLA reaction compliance
        if t.sla_reaction_deadline is not None:
            reaction_total += 1
            if t.sla_reaction_violated:
                reaction_violated += 1
        elif t.sla_deadline is not None:
            # fallback to old sla_deadline for tickets created before new SLA system
            reaction_total += 1
            if t.sla_deadline < now and t.status not in FINAL_STATUSES:
                reaction_violated += 1

        # SLA resolution compliance
        if t.sla_resolution_deadline is not None:
            resolution_total += 1
            if t.sla_resolution_violated:
                resolution_violated += 1
        elif t.sla_deadline is not None:
            resolution_total += 1
            if t.sla_deadline < now and t.status not in FINAL_STATUSES:
                resolution_violated += 1

        # Average resolution time (for closed/completed tickets)
        if t.status in ("closed", "completed") and t.closed_at and t.created_at:
            delta = (t.closed_at - t.created_at).total_seconds() / 3600
            resolution_times.append(delta)

    sla_reaction_pct = (
        None if reaction_total == 0
        else round(100 * (reaction_total - reaction_violated) / reaction_total, 2)
    )
    sla_resolution_pct = (
        None if resolution_total == 0
        else round(100 * (resolution_total - resolution_violated) / resolution_total, 2)
    )
    avg_resolution_hours = (
        None if not resolution_times
        else round(sum(resolution_times) / len(resolution_times), 1)
    )

    return {
        "total": total,
        "by_status": by_status,
        "by_type": by_type,
        "by_priority": by_priority,
        "by_engineer": by_engineer,
        "sla_reaction_compliance_pct": sla_reaction_pct,
        "sla_resolution_compliance_pct": sla_resolution_pct,
        "sla_compliance_pct": sla_resolution_pct,  # backward compat
        "avg_resolution_hours": avg_resolution_hours,
        "rows": rows,
        "period_from": date_from,
        "period_to": date_to,
    }


@router.get("/tickets", response_model=TicketReportResponse)
def report_tickets(
    date_from: date = Query(...),
    date_to: date = Query(...),
    engineer_id: Optional[int] = None,
    client_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(*_ROLES)),
):
    data = _build_report(db, date_from, date_to, engineer_id, client_id)
    return TicketReportResponse(
        total=data["total"],
        by_status=data["by_status"],
        by_type=data["by_type"],
        by_priority=data["by_priority"],
        by_engineer=data["by_engineer"],
        sla_reaction_compliance_pct=data["sla_reaction_compliance_pct"],
        sla_resolution_compliance_pct=data["sla_resolution_compliance_pct"],
        sla_compliance_pct=data["sla_compliance_pct"],
        avg_resolution_hours=data["avg_resolution_hours"],
        period_from=data["period_from"],
        period_to=data["period_to"],
    )


@router.get("/tickets/export/xlsx")
def export_tickets_xlsx(
    date_from: date = Query(...),
    date_to: date = Query(...),
    engineer_id: Optional[int] = None,
    client_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(*_ROLES)),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = _build_report(db, date_from, date_to, engineer_id, client_id)
    wb = Workbook()

    ws = wb.active
    ws.title = "Заявки"
    headers = ["Номер", "Создана", "Клиент", "Тип", "Приоритет", "Статус",
               "Инженер", "SLA реакция нарушена", "SLA решение нарушено"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for t in data["rows"]:
        ws.append([
            t.number,
            t.created_at.strftime("%d.%m.%Y %H:%M") if t.created_at else "",
            t.client.name if t.client else "",
            t.type,
            t.priority,
            t.status,
            t.assignee.full_name if t.assignee else "",
            "Да" if t.sla_reaction_violated else "Нет",
            "Да" if t.sla_resolution_violated else "Нет",
        ])

    ws2 = wb.create_sheet("Сводка")
    ws2.append(["Показатель", "Значение"])
    ws2.append(["Период с", str(date_from)])
    ws2.append(["Период по", str(date_to)])
    ws2.append(["Всего заявок", data["total"]])
    ws2.append(["% соблюдения SLA реакции",
                data["sla_reaction_compliance_pct"] if data["sla_reaction_compliance_pct"] is not None else "N/A"])
    ws2.append(["% соблюдения SLA решения",
                data["sla_resolution_compliance_pct"] if data["sla_resolution_compliance_pct"] is not None else "N/A"])
    ws2.append(["Среднее время решения (ч)",
                data["avg_resolution_hours"] if data["avg_resolution_hours"] is not None else "N/A"])
    ws2.append([])
    ws2.append(["По статусам"])
    for k, v in data["by_status"].items():
        ws2.append([k, v])
    ws2.append([])
    ws2.append(["По инженерам"])
    for k, v in data["by_engineer"].items():
        ws2.append([k, v])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"tickets_report_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
